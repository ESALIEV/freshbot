import io
import csv
from datetime import date

from aiogram import Router, F
from aiogram.types import Message, CallbackQuery, BufferedInputFile
from aiogram.filters import Command
from aiogram.fsm.context import FSMContext
from aiogram.utils.keyboard import InlineKeyboardBuilder

from db.database import (
    get_or_create_user, get_user_stores, get_store_stats,
    get_store_products_filtered, get_member_role,
)

router = Router()

MONTHS_RU = [
    "", "январе", "феврале", "марте", "апреле", "мае", "июне",
    "июле", "августе", "сентябре", "октябре", "ноябре", "декабре",
]


def current_month_ru() -> str:
    return MONTHS_RU[date.today().month]


def status_label(days_left: int) -> str:
    if days_left < 0:   return "❌ Просрочен"
    if days_left == 0:  return "🚨 Истекает сегодня"
    if days_left <= 3:  return "⚠️ Скоро истечёт"
    return "✅ Норма"


def filter_keyboard(store_id: int, active: str = ""):
    builder = InlineKeyboardBuilder()
    filters = [
        ("expired", "❌ Просроченные"),
        ("warning", "⚠️ Ближайшие 3 дня"),
        ("month",   f"📅 В {current_month_ru()}"),
        ("",        "📋 Все товары"),
    ]
    for key, label in filters:
        mark = "▶ " if key == active else ""
        builder.button(text=f"{mark}{label}", callback_data=f"filter:{store_id}:{key}")
    export_key = active if active else "all"
    builder.button(text="📥 Excel", callback_data=f"export:{store_id}:xlsx:{export_key}")
    builder.button(text="📥 CSV",   callback_data=f"export:{store_id}:csv:{export_key}")
    builder.adjust(1, 1, 1, 1, 2)
    return builder.as_markup()


# ── /stats ──

@router.message(Command("stats"))
@router.message(F.text == "📊 Статистика")
async def cmd_stats(message: Message, state: FSMContext):
    user = await get_or_create_user(message.from_user.id)
    stores = await get_user_stores(user["id"])

    if not stores:
        await message.answer("У вас нет магазинов. /newstore или /join")
        return

    if len(stores) == 1:
        await _send_stats(message, stores[0]["id"], stores[0]["name"])
    else:
        builder = InlineKeyboardBuilder()
        for s in stores:
            builder.button(text=f"🏪 {s['name']}", callback_data=f"stats_store:{s['id']}")
        builder.adjust(1)
        await message.answer("📊 Статистика какого магазина?", reply_markup=builder.as_markup())


@router.callback_query(F.data.startswith("stats_store:"))
async def cb_stats_store(callback: CallbackQuery, state: FSMContext):
    store_id = int(callback.data.split(":")[1])
    user = await get_or_create_user(callback.from_user.id)
    if not await get_member_role(user["id"], store_id):
        await callback.answer("❌ Нет доступа", show_alert=True)
        return
    stores = await get_user_stores(user["id"])
    store  = next((s for s in stores if s["id"] == store_id), None)
    await _send_stats(callback.message, store_id, store["name"] if store else "")
    await callback.answer()


async def _send_stats(message: Message, store_id: int, store_name: str):
    s = await get_store_stats(store_id)
    month = current_month_ru()

    total         = s.get("total")              or 0
    expired       = s.get("expired")            or 0
    expires_3d    = s.get("expires_3d")         or 0
    expires_month = s.get("expires_this_month") or 0
    total_qty     = s.get("total_qty")          or 0

    text = (
        f"📊 <b>Статистика — {store_name}</b>\n\n"
        f"📦 Позиций всего: <b>{total}</b> ({total_qty} шт.)\n"
        f"❌ Просрочено: <b>{expired}</b>\n"
        f"⚠️ Истекают в ближайшие 3 дня: <b>{expires_3d}</b>\n"
        f"📅 Истекут в {month}: <b>{expires_month}</b>"
    )

    builder = InlineKeyboardBuilder()
    if expired:
        builder.button(text=f"❌ Просроченные ({expired})",        callback_data=f"filter:{store_id}:expired")
    if expires_3d:
        builder.button(text=f"⚠️ Ближайшие 3 дня ({expires_3d})", callback_data=f"filter:{store_id}:warning")
    if expires_month:
        builder.button(text=f"📅 В {month} ({expires_month})",     callback_data=f"filter:{store_id}:month")
    builder.button(text="📥 Экспорт Excel", callback_data=f"export:{store_id}:xlsx:all")
    builder.button(text="📥 Экспорт CSV",   callback_data=f"export:{store_id}:csv:all")
    builder.adjust(1)

    await message.answer(text, reply_markup=builder.as_markup(), parse_mode="HTML")


# ── Фильтры ──

FILTER_HEADERS = {
    "expired": "❌ Просроченные товары",
    "warning": "⚠️ Истекают в ближайшие 3 дня",
    "month":   lambda: f"📅 Истекают в {current_month_ru()}",
    "":        "📋 Все товары",
}


@router.callback_query(F.data.startswith("filter:"))
async def cb_filter(callback: CallbackQuery):
    _, store_id_str, status = callback.data.split(":", 2)
    store_id = int(store_id_str)

    user = await get_or_create_user(callback.from_user.id)
    if not await get_member_role(user["id"], store_id):
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    products = await get_store_products_filtered(store_id, status)
    header_val = FILTER_HEADERS.get(status, "Товары")
    header = header_val() if callable(header_val) else header_val
    nav = filter_keyboard(store_id, status)

    if not products:
        await callback.message.answer(f"{header}\n\n✅ Таких товаров нет.", reply_markup=nav)
        await callback.answer()
        return

    await callback.message.answer(f"{header} — {len(products)} шт.", reply_markup=nav)

    for p in products:
        days = p["days_left"]
        emoji = status_label(days).split()[0]
        if days < 0:      days_text = f"просрочен {abs(days)} дн. назад"
        elif days == 0:   days_text = "истекает СЕГОДНЯ"
        elif days == 1:   days_text = "завтра"
        else:             days_text = f"через {days} дн."

        article = f"\nАртикул: <code>{p['article']}</code>" if p.get("article") else ""
        text = (
            f"{emoji} <b>{p['name']}</b>{article}\n"
            f"Кол-во: {p['quantity']} шт. · Срок: {p['expiry_date']} ({days_text})"
        )
        row = InlineKeyboardBuilder()
        row.button(text="✏️", callback_data=f"edit:{p['batch_id']}")
        row.button(text="🗑️", callback_data=f"delete:{p['batch_id']}")
        await callback.message.answer(text, reply_markup=row.as_markup(), parse_mode="HTML")

    await callback.answer()


# ── Экспорт ──

FILTER_NAMES = {
    "all":     "все товары",
    "expired": "просроченные",
    "warning": "истекают 3д",
    "month":   lambda: f"истекают в {current_month_ru()}",
}

ROW_COLORS = {
    "❌ Просрочен":        "FFCDD2",
    "🚨 Истекает сегодня": "FFE0B2",
    "⚠️ Скоро истечёт":   "FFF9C4",
    "✅ Норма":            "E8F5E9",
}


@router.callback_query(F.data.startswith("export:"))
async def cb_export(callback: CallbackQuery):
    _, store_id_str, fmt, status = callback.data.split(":", 3)
    store_id = int(store_id_str)

    user = await get_or_create_user(callback.from_user.id)
    if not await get_member_role(user["id"], store_id):
        await callback.answer("❌ Нет доступа", show_alert=True)
        return

    await callback.answer("⏳ Готовлю файл...")

    filter_key   = "" if status == "all" else status
    products     = await get_store_products_filtered(store_id, filter_key)

    if not products:
        await callback.message.answer("📭 Нет товаров для экспорта.")
        return

    stores       = await get_user_stores(user["id"])
    store        = next((s for s in stores if s["id"] == store_id), None)
    store_name   = store["name"] if store else f"store_{store_id}"
    date_str     = date.today().strftime("%d.%m.%Y")
    date_file    = date.today().strftime("%Y%m%d")
    fn_val       = FILTER_NAMES.get(status, status)
    filter_label = fn_val() if callable(fn_val) else fn_val

    if fmt == "xlsx":
        file_bytes = _make_xlsx(products, store_name, date_str, filter_label)
        filename   = f"freshbot_{store_name}_{status}_{date_file}.xlsx"
    else:
        file_bytes = _make_csv(products)
        filename   = f"freshbot_{store_name}_{status}_{date_file}.csv"

    await callback.message.answer_document(
        BufferedInputFile(file_bytes, filename=filename),
        caption=(
            f"📥 <b>{store_name}</b> · {filter_label}\n"
            f"Позиций: {len(products)} · {date_str}"
        ),
        parse_mode="HTML",
    )


def _make_xlsx(products: list, store_name: str, date_str: str, filter_label: str) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Товары"

    ws.merge_cells("A1:F1")
    tc = ws["A1"]
    tc.value     = f"{store_name} — {filter_label} · {date_str}"
    tc.font      = Font(bold=True, size=13)
    tc.alignment = Alignment(horizontal="center")

    headers     = ["№", "Название", "Артикул", "Кол-во", "Срок годности", "Статус"]
    header_fill = PatternFill("solid", fgColor="2E7D32")
    header_font = Font(bold=True, color="FFFFFF")
    thin        = Side(style="thin")
    border      = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(headers, 1):
        cell           = ws.cell(row=2, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border    = border

    for i, p in enumerate(products, 1):
        slabel   = status_label(p["days_left"])
        row_data = [i, p["name"], p.get("article") or "—", p["quantity"], p["expiry_date"], slabel]
        fill     = PatternFill("solid", fgColor=ROW_COLORS.get(slabel, "FFFFFF"))
        for col, val in enumerate(row_data, 1):
            cell           = ws.cell(row=i + 2, column=col, value=val)
            cell.fill      = fill
            cell.border    = border
            cell.alignment = Alignment(horizontal="center" if col in (1, 4) else "left")

    for col, w in enumerate([5, 35, 15, 10, 18, 22], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    ws.freeze_panes = "A3"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _make_csv(products: list) -> bytes:
    buf = io.StringIO()
    writer = csv.writer(buf, delimiter=";")
    writer.writerow(["№", "Название", "Артикул", "Кол-во", "Срок годности", "Статус", "Дней осталось"])
    for i, p in enumerate(products, 1):
        writer.writerow([
            i,
            p["name"],
            p.get("article") or "",
            p["quantity"],
            p["expiry_date"],
            status_label(p["days_left"]).replace("❌ ", "").replace("🚨 ", "").replace("⚠️ ", "").replace("✅ ", ""),
            p["days_left"],
        ])
    return buf.getvalue().encode("utf-8-sig")
